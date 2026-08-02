from __future__ import annotations

import calendar
import hashlib
import json
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parents[2]
APP = ROOT / "Breakfast"
ACCOUNTING_OUTPUT = APP / "accounting" / "data" / "revenue-history-2022-2025.js"
SALARY_OUTPUT = APP / "salary_app" / "data" / "salary-history-2022-2025.js"
REPORT_OUTPUT = APP / "data_import_reports" / "history-2022-2025-report.json"

# Historical workbooks sometimes use a nickname or put an accounting-only
# balancing label in the employee column. Keep the rules explicit so the
# generated history is reproducible and those labels never become employees.
EMPLOYEE_NAME_ALIASES = {"采葳": "黃采葳"}
EXCLUDED_PAYROLL_NAMES = {"年終", "待補", "其他薪資支出（原營業額檔）"}

OWNERS = {"以馨", "月霞"}
FIXED_COST = {"房租", "電費", "水費", "瓦斯", "瓦斯費", "電話費", "稅金/保險", "平台", "平台手續費"}
DRINK_COST = {"開元", "聖淘沙", "冰塊", "上統茶葉"}
GENERAL_COST = {"寶綠", "寶綠餐具", "大燈籠", "萬霖", "十全"}


def scalar(value: Any) -> float:
    if isinstance(value, bool) or value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return 0.0


def clean_amount(value: Any) -> int | float:
    number = scalar(value)
    return int(round(number)) if abs(number - round(number)) < 1e-8 else round(number, 6)


def text(value: Any) -> str:
    return str(value).strip() if value not in (None, "") else ""


def source_date(value: Any, year: int, month: int, day_hint: int | None = None) -> str | None:
    parsed: date | None = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, (int, float)):
        if 1 <= value <= 31 and day_hint is None:
            parsed = date(year, month, int(value))
        elif value > 30000:
            converted = from_excel(value)
            parsed = converted.date() if isinstance(converted, datetime) else converted
    elif text(value):
        raw = text(value)
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d"):
            try:
                candidate = datetime.strptime(raw, fmt)
                parsed = candidate.date() if "%Y" in fmt else date(year, candidate.month, candidate.day)
                break
            except ValueError:
                pass
    if not parsed and day_hint and day_hint <= calendar.monthrange(year, month)[1]:
        parsed = date(year, month, day_hint)
    return parsed.isoformat() if parsed and parsed.year == year and parsed.month == month else None


def clock(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)) and 0 <= value < 1:
        minutes = round(value * 24 * 60) % (24 * 60)
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    raw = text(value)
    match = re.search(r"(\d{1,2})[:：](\d{2})", raw)
    if match:
        hour, minute = int(match.group(1)), int(match.group(2))
        if 0 <= hour < 24 and 0 <= minute < 60:
            return f"{hour:02d}:{minute:02d}"
    return None


def minutes_between(start: str | None, end: str | None) -> int:
    if not start or not end:
        return 0
    sh, sm = map(int, start.split(":"))
    eh, em = map(int, end.split(":"))
    result = eh * 60 + em - sh * 60 - sm
    return result + 24 * 60 if result < 0 else result


def stable_id(*parts: Any) -> str:
    digest = hashlib.sha1("|".join(map(str, parts)).encode("utf-8")).hexdigest()[:14]
    return f"history-{digest}"


def employee_id(name: str) -> str:
    return f"history-employee-{hashlib.sha1(name.encode('utf-8')).hexdigest()[:10]}"


def canonical_employee_name(name: Any) -> str:
    cleaned = text(name)
    return EMPLOYEE_NAME_ALIASES.get(cleaned, cleaned)


def cell(rows: list[tuple[Any, ...]], row: int, column: int) -> Any:
    if row < 1 or row > len(rows):
        return None
    values = rows[row - 1]
    return values[column - 1] if 1 <= column <= len(values) else None


def cached_rows(sheet, max_row: int, max_col: int) -> list[tuple[Any, ...]]:
    return list(sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True))


def expense_group(category: str) -> str:
    if category in FIXED_COST or any(word in category for word in ("房租", "電費", "水費", "瓦斯", "電話", "稅金", "保險", "平台手續")):
        return "固定成本"
    if category in DRINK_COST or any(word in category for word in ("茶葉", "飲料", "冰塊")):
        return "飲品成本"
    if category in GENERAL_COST or any(word in category for word in ("餐具", "清潔", "雜貨")):
        return "雜貨成本"
    if "雜支" in category:
        return "雜支"
    return "食材成本"


def income_group(category: str) -> str:
    lowered = category.lower()
    if "uber" in lowered or "熊貓" in category or "foodpanda" in lowered:
        return "平台收入"
    if "其他" in category:
        return "其他收入"
    return "現金收入"


def make_transaction(*, year: int, month: int, source_name: str, sheet_name: str, source_ref: str,
                     tx_date: str, tx_type: str, group: str, category: str, amount: Any) -> dict[str, Any]:
    method = "平台入帳" if tx_type == "income" and group == "平台收入" else ("營業收入" if tx_type == "income" else "營業支出")
    return {
        "id": stable_id(year, month, sheet_name, source_ref, tx_date, tx_type, group, category),
        "date": tx_date,
        "type": tx_type,
        "group": group,
        "category": category,
        "amount": clean_amount(amount),
        "paymentMethod": method,
        "counterparty": category,
        "note": f"由 {year} 年營業額活頁簿匯入",
        "source": "workbook-history",
        "sourceRef": f"{source_name}・{sheet_name}・{source_ref}",
        "locked": True,
    }


def parse_new_accounting(book, year: int, month: int, source_name: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    sheet_name = f"{month}月營業表"
    sheet = book[sheet_name]
    rows = cached_rows(sheet, 55, 70)
    date_columns: list[tuple[int, str]] = []
    for column in range(7, 68, 2):
        tx_date = source_date(cell(rows, 3, column), year, month, (column - 7) // 2 + 1)
        if tx_date:
            date_columns.append((column, tx_date))

    income_total_row = next(row for row in range(5, 20) if text(cell(rows, row, 5)) == "本日收入")
    expense_total_row = next(row for row in range(income_total_row + 2, 55) if text(cell(rows, row, 5)) == "本日支出")
    transactions: list[dict[str, Any]] = []

    group = ""
    for row in range(5, income_total_row):
        group = text(cell(rows, row, 5)) or group
        category = text(cell(rows, row, 6))
        if not category:
            continue
        resolved_group = group if "收入" in group else income_group(category)
        for column, tx_date in date_columns:
            amount = scalar(cell(rows, row, column))
            if amount:
                transactions.append(make_transaction(
                    year=year, month=month, source_name=source_name, sheet_name=sheet_name,
                    source_ref=f"第 {row} 列・{tx_date}", tx_date=tx_date, tx_type="income",
                    group=resolved_group, category=category, amount=amount,
                ))

    payroll_source = 0.0
    group = ""
    for row in range(income_total_row + 2, expense_total_row):
        group = text(cell(rows, row, 5)) or group
        category = text(cell(rows, row, 6))
        if not category:
            continue
        row_total = scalar(cell(rows, row, 69))
        if group == "人事成本" or category == "薪資":
            payroll_source += row_total
            continue
        resolved_group = group if group and group != "人事成本" else expense_group(category)
        for column, tx_date in date_columns:
            amount = scalar(cell(rows, row, column))
            if amount:
                transactions.append(make_transaction(
                    year=year, month=month, source_name=source_name, sheet_name=sheet_name,
                    source_ref=f"第 {row} 列・{tx_date}", tx_date=tx_date, tx_type="expense",
                    group=resolved_group, category=category, amount=amount,
                ))

    expected_income = scalar(cell(rows, income_total_row, 69))
    expected_expense = scalar(cell(rows, expense_total_row, 69))
    imported_income = sum(scalar(item["amount"]) for item in transactions if item["type"] == "income")
    imported_operating = sum(scalar(item["amount"]) for item in transactions if item["type"] == "expense")
    reconciliation = {
        "period": f"{year}-{month:02d}",
        "sourceSheet": sheet_name,
        "expectedIncome": clean_amount(expected_income),
        "importedIncome": clean_amount(imported_income),
        "incomeDelta": clean_amount(imported_income - expected_income),
        "expectedOperating": clean_amount(expected_expense - payroll_source),
        "importedOperating": clean_amount(imported_operating),
        "expenseDelta": clean_amount(imported_operating - (expected_expense - payroll_source)),
        "sourcePayroll": clean_amount(payroll_source),
    }
    return transactions, reconciliation


def old_2022_file(month: int) -> Path:
    old = ROOT / "舊資料"
    matches = list(old.glob(f"*營業額(2022年{month}月)*.xlsx"))
    if not matches:
        raise FileNotFoundError(f"找不到 2022 年 {month} 月舊營業額檔")
    return matches[0]


def parse_old_2022_accounting(month: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    path = old_2022_file(month)
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    income_sheet = next(sheet for sheet in book.worksheets if "收入統計" in sheet.title)
    expense_sheet = next(sheet for sheet in book.worksheets if "支出統計" in sheet.title)
    income_rows = cached_rows(income_sheet, 60, 12)
    expense_rows = cached_rows(expense_sheet, 50, min(expense_sheet.max_column, 40))
    transactions: list[dict[str, Any]] = []

    income_headers = {column: text(cell(income_rows, 3, column)) for column in range(3, 9)}
    income_total_row = 0
    for row in range(4, 60):
        if text(cell(income_rows, row, 1)) == "小計":
            income_total_row = row
            break
        tx_date = source_date(cell(income_rows, row, 1), 2022, month)
        if not tx_date:
            continue
        for column, category in income_headers.items():
            amount = scalar(cell(income_rows, row, column))
            if amount and category:
                transactions.append(make_transaction(
                    year=2022, month=month, source_name=path.name, sheet_name=income_sheet.title,
                    source_ref=f"第 {row} 列・{tx_date}", tx_date=tx_date, tx_type="income",
                    group=income_group(category), category=category, amount=amount,
                ))

    expense_headers = {column: text(cell(expense_rows, 3, column)) for column in range(3, len(expense_rows[2]))}
    expense_total_row = 0
    payroll_source = 0.0
    for row in range(4, 50):
        if text(cell(expense_rows, row, 1)) == "小計":
            expense_total_row = row
            break
        tx_date = source_date(cell(expense_rows, row, 1), 2022, month)
        if not tx_date:
            continue
        for column, category in expense_headers.items():
            amount = scalar(cell(expense_rows, row, column))
            if not amount or not category or category == "當天總支出":
                continue
            if category == "薪資":
                payroll_source += amount
                continue
            transactions.append(make_transaction(
                year=2022, month=month, source_name=path.name, sheet_name=expense_sheet.title,
                source_ref=f"第 {row} 列・{tx_date}", tx_date=tx_date, tx_type="expense",
                group=expense_group(category), category=category, amount=amount,
            ))

    expected_income = scalar(cell(income_rows, income_total_row, 9))
    expected_expense = scalar(cell(expense_rows, expense_total_row, len(expense_rows[2])))
    if not payroll_source:
        payroll_column = next((column for column, category in expense_headers.items() if category == "薪資"), 0)
        payroll_source = scalar(cell(expense_rows, expense_total_row, payroll_column)) if payroll_column else 0
    imported_income = sum(scalar(item["amount"]) for item in transactions if item["type"] == "income")
    imported_operating = sum(scalar(item["amount"]) for item in transactions if item["type"] == "expense")
    payroll_details: list[dict[str, Any]] = []
    for row in range(income_total_row + 1, min(len(income_rows), income_total_row + 25)):
        for column in range(8, min(len(income_rows[row - 1]), 12) + 1):
            label = text(cell(income_rows, row, column))
            amount = scalar(cell(income_rows, row + 1, column))
            if label and amount and label not in {"薪資", "占比", "人事成本"}:
                payroll_details.append({"name": label, "amount": clean_amount(amount), "sourceCell": f"{income_sheet.title} {row + 1}:{column}"})
    if abs(sum(scalar(item["amount"]) for item in payroll_details) - payroll_source) > 1:
        payroll_details = []
    book.close()
    return transactions, {
        "period": f"2022-{month:02d}",
        "sourceSheet": f"{income_sheet.title}＋{expense_sheet.title}",
        "expectedIncome": clean_amount(expected_income),
        "importedIncome": clean_amount(imported_income),
        "incomeDelta": clean_amount(imported_income - expected_income),
        "expectedOperating": clean_amount(expected_expense - payroll_source),
        "importedOperating": clean_amount(imported_operating),
        "expenseDelta": clean_amount(imported_operating - (expected_expense - payroll_source)),
        "sourcePayroll": clean_amount(payroll_source),
        "sourcePayrollDetails": payroll_details,
    }


def scan_rate_profile(rows: list[tuple[Any, ...]], start_col: int, end_col: int, row_start: int, row_end: int,
                      name: str, effective_from: str, schedule_start: str | None = None) -> dict[str, Any]:
    hourly = weekend = holiday = monthly = 0.0
    schedule_end: str | None = None
    for row in range(row_start, row_end + 1):
        for column in range(start_col, end_col + 1):
            label = text(cell(rows, row, column))
            if not label:
                continue
            numerics = [scalar(cell(rows, row, target)) for target in range(column + 1, end_col + 1)]
            value = next((number for number in numerics if number), 0)
            if "基本薪資" in label:
                monthly = value
            elif "假日時薪" in label:
                weekend = value
            elif "平日時薪" in label:
                hourly = value
            elif label == "時薪" or label.endswith("時薪"):
                hourly = value
    if name in OWNERS:
        monthly = monthly or 1
    pay_type = "monthly" if monthly else "hourly"
    if pay_type == "hourly":
        hourly = hourly or weekend or holiday
        weekend = weekend or hourly
        holiday = holiday or hourly
    profile: dict[str, Any] = {
        "name": name,
        "id": employee_id(name),
        "payType": pay_type,
        "effectiveFrom": effective_from,
        "active": False,
    }
    if pay_type == "monthly":
        profile.update({
            "monthlySalary": clean_amount(monthly if monthly > 1 else 0),
            "scheduleStart": schedule_start or "",
            "scheduleEnd": schedule_end or "",
            "attendanceRequired": name not in OWNERS,
            "overtimeMode": "none",
            "overtimeHourlyRate": 0,
        })
    else:
        profile.update({
            "hourlyRate": clean_amount(hourly),
            "weekendRate": clean_amount(weekend),
            "holidayRate": clean_amount(holiday),
        })
    return profile


def payroll_row(name: str, total: Any, pay_type: str, minutes: int, source_cell: str) -> dict[str, Any]:
    amount = clean_amount(total)
    return {
        "employeeName": name,
        "payType": pay_type,
        "regularPay": amount,
        "overtimePay": 0,
        "specialPay": 0,
        "earnings": 0,
        "deductions": 0,
        "total": amount,
        "regularMinutes": minutes,
        "overtimeMinutes": 0,
        "issues": 0,
        "detailLines": [{"label": "原工資簿實領薪資", "amount": amount}],
        "adjustments": [],
        "sourceCell": source_cell,
    }


def parse_new_salary(book, year: int, month: int, source_name: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    sheet_name = f"{month}月薪資"
    sheet = book[sheet_name]
    max_col = min(max(sheet.max_column, 80), 180)
    rows = cached_rows(sheet, 75, max_col)
    summary: list[tuple[str, float, str]] = []
    for row in range(5, 40):
        name = text(cell(rows, row, 2))
        if not name or "本月人事成本" in name:
            if "本月人事成本" in name:
                break
            continue
        total = scalar(cell(rows, row, 3))
        if total:
            summary.append((name, total, f"{sheet_name} C{row}"))

    summary_names = {name for name, _, _ in summary}
    block_starts = [column for column in range(1, max_col + 1) if text(cell(rows, 3, column)) in summary_names]
    block_by_name: dict[str, tuple[int, int]] = {}
    for index, start_col in enumerate(block_starts):
        name = text(cell(rows, 3, start_col))
        end_col = block_starts[index + 1] - 1 if index + 1 < len(block_starts) else min(start_col + 9, max_col)
        block_by_name[name] = (start_col, end_col)

    attendance: list[dict[str, Any]] = []
    employees: list[dict[str, Any]] = []
    payroll: list[dict[str, Any]] = []
    for name, total, source_cell in summary:
        start_col, end_col = block_by_name.get(name, (0, 0))
        total_minutes = 0
        schedule_start = None
        if start_col:
            schedule_start = clock(cell(rows, 5, start_col + 2))
            for row in range(5, 36):
                tx_date = source_date(cell(rows, row, start_col), year, month)
                start = clock(cell(rows, row, start_col + 3))
                end = clock(cell(rows, row, start_col + 4))
                if not tx_date or not start or not end:
                    continue
                source_minutes = int(round(scalar(cell(rows, row, start_col + 6)))) or minutes_between(start, end)
                total_minutes += source_minutes
                attendance.append({
                    "employeeName": name,
                    "date": tx_date,
                    "segments": [{"start": start, "end": end}],
                    "sourceMinutes": source_minutes,
                    "roundedMinutes": source_minutes,
                    "sourceDayPay": clean_amount(cell(rows, row, start_col + 7)),
                    "sourceCells": f"{sheet_name} R{row}C{start_col}:C{start_col + 7}",
                })
            profile = scan_rate_profile(rows, start_col, end_col, 36, 70, name, f"{year}-{month:02d}-01", schedule_start)
        else:
            profile = scan_rate_profile(rows, 1, 5, 36, 70, name, f"{year}-{month:02d}-01")
            profile.update({"payType": "monthly", "monthlySalary": clean_amount(total), "attendanceRequired": name not in OWNERS, "overtimeMode": "none", "overtimeHourlyRate": 0})
        if profile["payType"] == "monthly" and not profile.get("monthlySalary"):
            profile["monthlySalary"] = clean_amount(total)
        employees.append(profile)
        payroll.append(payroll_row(name, total, profile["payType"], total_minutes, source_cell))
    return employees, attendance, payroll


def parse_old_2022_salary(book, month: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    sheet = book.worksheets[11 - month]
    rows = cached_rows(sheet, 70, min(sheet.max_column, 160))
    starts = [column for column in range(1, len(rows[0]) + 1) if text(cell(rows, 1, column))]
    raw_names = [text(cell(rows, 2, column + 3)) for column in starts]
    occurrences: dict[str, int] = {}
    attendance: list[dict[str, Any]] = []
    employees: list[dict[str, Any]] = []
    payroll: list[dict[str, Any]] = []
    for index, start_col in enumerate(starts):
        end_col = (starts[index + 1] - 1) if index + 1 < len(starts) else min(start_col + 9, len(rows[1]))
        original_name = raw_names[index]
        occurrences[original_name] = occurrences.get(original_name, 0) + 1
        name = original_name if occurrences[original_name] == 1 else f"{original_name}（來源第 {occurrences[original_name]} 欄）"
        total = 0.0
        total_source = ""
        for row in range(35, 70):
            labels = " ".join(text(cell(rows, row, column)) for column in range(start_col, end_col + 1))
            if "合計薪資" in labels or "實領薪資" in labels:
                numerics = [scalar(cell(rows, row, column)) for column in range(start_col, end_col + 1)]
                total = next((number for number in reversed(numerics) if number), 0)
                total_source = f"{sheet.title} R{row}C{start_col}:C{end_col}"
        if not total:
            continue
        total_minutes = 0
        schedule_start = clock(cell(rows, 4, start_col + 2))
        for row in range(4, 35):
            tx_date = source_date(cell(rows, row, start_col), 2022, month)
            start = clock(cell(rows, row, start_col + 3))
            end = clock(cell(rows, row, start_col + 4))
            if not tx_date or not start or not end:
                continue
            source_minutes = int(round(scalar(cell(rows, row, start_col + 6)))) or minutes_between(start, end)
            total_minutes += source_minutes
            attendance.append({
                "employeeName": name,
                "date": tx_date,
                "segments": [{"start": start, "end": end}],
                "sourceMinutes": source_minutes,
                "roundedMinutes": source_minutes,
                "sourceDayPay": clean_amount(cell(rows, row, start_col + 7)),
                "sourceCells": f"{sheet.title} R{row}C{start_col}:C{start_col + 7}",
            })
        profile = scan_rate_profile(rows, start_col, end_col, 35, 69, name, f"2022-{month:02d}-01", schedule_start)
        if profile["payType"] == "monthly" and not profile.get("monthlySalary"):
            profile["monthlySalary"] = clean_amount(total)
        employees.append(profile)
        payroll.append(payroll_row(name, total, profile["payType"], total_minutes, total_source))
    return employees, attendance, payroll


def workbook_for_year(year: int) -> Path:
    matches = list(ROOT.glob(f"初一食午營業額新版({year}年).xlsm"))
    if not matches:
        raise FileNotFoundError(f"找不到 {year} 年營業額新版")
    return matches[0]


def write_js(path: Path, global_name: str, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content = f"window.{global_name} = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n"
    path.write_text(content, encoding="utf-8", newline="\n")


def main() -> None:
    all_transactions: list[dict[str, Any]] = []
    all_reconciliation: list[dict[str, Any]] = []
    all_employees: list[dict[str, Any]] = []
    all_attendance: list[dict[str, Any]] = []
    all_payroll: dict[str, list[dict[str, Any]]] = {}

    old_salary_path = ROOT / "舊資料" / "初一食午薪資計算.xlsx"
    old_salary = openpyxl.load_workbook(old_salary_path, read_only=True, data_only=True)

    for month in range(1, 11):
        transactions, reconciliation = parse_old_2022_accounting(month)
        employees, attendance, wage_book_payroll = parse_old_2022_salary(old_salary, month)
        period = f"2022-{month:02d}"
        payroll = []
        profile_by_name = {profile["name"]: profile for profile in employees}
        alias = {"小亘": "林辰亘", "宥臻": "周宥臻", "丁丁": "丁翊盛"}
        for detail in reconciliation.get("sourcePayrollDetails", []):
            name = alias.get(detail["name"], detail["name"])
            profile = profile_by_name.get(name)
            if not profile:
                profile = {
                    "name": name,
                    "id": employee_id(name),
                    "payType": "monthly",
                    "monthlySalary": clean_amount(detail["amount"]),
                    "scheduleStart": "",
                    "scheduleEnd": "",
                    "attendanceRequired": name not in OWNERS,
                    "overtimeMode": "none",
                    "overtimeHourlyRate": 0,
                    "effectiveFrom": f"2022-{month:02d}-01",
                    "active": False,
                }
                employees.append(profile)
                profile_by_name[name] = profile
            payroll.append(payroll_row(name, detail["amount"], profile["payType"], 0, detail["sourceCell"]))
        if not payroll:
            payroll = wage_book_payroll
        reconciliation["salarySnapshotTotal"] = clean_amount(sum(scalar(row["total"]) for row in payroll))
        reconciliation["payrollDelta"] = clean_amount(reconciliation["salarySnapshotTotal"] - scalar(reconciliation["sourcePayroll"]))
        all_transactions.extend(transactions)
        all_reconciliation.append(reconciliation)
        all_employees.extend(employees)
        all_attendance.extend(attendance)
        all_payroll[period] = payroll
    old_salary.close()

    for year in range(2022, 2026):
        path = workbook_for_year(year)
        book = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        first_month = 11 if year == 2022 else 1
        for month in range(first_month, 13):
            transactions, reconciliation = parse_new_accounting(book, year, month, path.name)
            employees, attendance, payroll = parse_new_salary(book, year, month, path.name)
            period = f"{year}-{month:02d}"
            reconciliation["salarySnapshotTotal"] = clean_amount(sum(scalar(row["total"]) for row in payroll))
            reconciliation["payrollDelta"] = clean_amount(reconciliation["salarySnapshotTotal"] - scalar(reconciliation["sourcePayroll"]))
            all_transactions.extend(transactions)
            all_reconciliation.append(reconciliation)
            all_employees.extend(employees)
            all_attendance.extend(attendance)
            all_payroll[period] = payroll
        book.close()

    # Canonicalize every parser path at the final boundary. This makes the
    # nickname and full name share one employee id and keeps accounting-only
    # labels out of employees, attendance and payroll snapshots.
    cleaned_employees: list[dict[str, Any]] = []
    for profile in all_employees:
        name = canonical_employee_name(profile.get("name"))
        if not name or name in EXCLUDED_PAYROLL_NAMES:
            continue
        cleaned_employees.append({**profile, "name": name, "id": employee_id(name)})
    all_employees = cleaned_employees

    cleaned_attendance: list[dict[str, Any]] = []
    seen_attendance: set[tuple[Any, ...]] = set()
    for row in all_attendance:
        name = canonical_employee_name(row.get("employeeName"))
        if not name or name in EXCLUDED_PAYROLL_NAMES:
            continue
        normalized_row = {**row, "employeeName": name}
        signature = (
            name,
            normalized_row.get("date"),
            json.dumps(normalized_row.get("segments") or [], ensure_ascii=False, sort_keys=True),
            normalized_row.get("sourceCells"),
        )
        if signature in seen_attendance:
            continue
        seen_attendance.add(signature)
        cleaned_attendance.append(normalized_row)
    all_attendance = cleaned_attendance

    for period, rows in list(all_payroll.items()):
        merged: dict[str, dict[str, Any]] = {}
        for row in rows:
            name = canonical_employee_name(row.get("employeeName"))
            if not name or name in EXCLUDED_PAYROLL_NAMES:
                continue
            normalized_row = {**row, "employeeName": name}
            if name not in merged:
                merged[name] = normalized_row
                continue
            target = merged[name]
            for field in ("total", "baseSalary", "regularPay", "overtimePay", "bonus", "deduction", "totalMinutes"):
                target[field] = clean_amount(scalar(target.get(field)) + scalar(normalized_row.get(field)))
            target["sourceCell"] = "；".join(filter(None, [text(target.get("sourceCell")), text(normalized_row.get("sourceCell"))]))
            target["detailLines"] = [*(target.get("detailLines") or []), *(normalized_row.get("detailLines") or [])]
            target["adjustments"] = [*(target.get("adjustments") or []), *(normalized_row.get("adjustments") or [])]
        all_payroll[period] = sorted(merged.values(), key=lambda item: item["employeeName"])

    # Recalculate after removing erroneous labels. Any difference remains in
    # the import report instead of being silently assigned to a real employee.
    for reconciliation in all_reconciliation:
        period = reconciliation["period"]
        snapshot_total = sum(scalar(row["total"]) for row in all_payroll.get(period, []))
        reconciliation["salarySnapshotTotal"] = clean_amount(snapshot_total)
        reconciliation["payrollDelta"] = clean_amount(snapshot_total - scalar(reconciliation["sourcePayroll"]))

    unique_profiles: dict[tuple[str, str], dict[str, Any]] = {}
    for profile in all_employees:
        unique_profiles[(profile["name"], profile["effectiveFrom"])] = profile
    all_employees = sorted(unique_profiles.values(), key=lambda item: (item["name"], item["effectiveFrom"]))
    all_attendance.sort(key=lambda item: (item["date"], item["employeeName"]))
    all_transactions.sort(key=lambda item: (item["date"], item["type"], item["group"], item["category"], item["id"]))
    all_reconciliation.sort(key=lambda item: item["period"])

    accounting_payload = {
        "id": "revenue-workbooks-2022-2025-v1",
        "source": "初一食午營業額新版（2022–2025 年；2022 年 1–10 月補舊資料）",
        "importedThrough": "2025-12-31",
        "generatedAt": date.today().isoformat(),
        "policy": {
            "payroll": "正式員工薪資排除於記帳支出，改由薪資管理歷史快照串接，避免重複計算。",
            "legacy2022": "2022 年新版僅含 11–12 月，1–10 月採舊資料夾中的逐月營業額檔。",
            "locked": "來源活頁簿交易為鎖定歷史資料，本機新增項目不受影響。",
        },
        "transactions": all_transactions,
        "reconciliation": all_reconciliation,
    }
    salary_payload = {
        "id": "salary-workbooks-2022-2025-v1",
        "source": "2022 年舊薪資簿（1–10 月）＋2022–2025 年營業額新版內嵌薪資表",
        "sourceMonths": sorted(all_payroll),
        "employees": all_employees,
        "attendance": all_attendance,
        "payroll": all_payroll,
    }
    write_js(ACCOUNTING_OUTPUT, "BREAKFAST_ACCOUNTING_HISTORY_2022_2025", accounting_payload)
    write_js(SALARY_OUTPUT, "BREAKFAST_SALARY_HISTORY_2022_2025", salary_payload)

    monthly_report = []
    for item in all_reconciliation:
        period = item["period"]
        monthly_report.append({
            **item,
            "transactionCount": sum(1 for tx in all_transactions if tx["date"].startswith(period)),
            "attendanceCount": sum(1 for row in all_attendance if row["date"].startswith(period)),
            "payrollEmployeeCount": len(all_payroll.get(period, [])),
        })
    report = {
        "generatedAt": date.today().isoformat(),
        "period": "2022-01～2025-12",
        "totals": {
            "transactions": len(all_transactions),
            "attendance": len(all_attendance),
            "payrollMonths": len(all_payroll),
            "payrollRows": sum(len(rows) for rows in all_payroll.values()),
            "income": clean_amount(sum(scalar(tx["amount"]) for tx in all_transactions if tx["type"] == "income")),
            "operatingExpenseExcludingPayroll": clean_amount(sum(scalar(tx["amount"]) for tx in all_transactions if tx["type"] == "expense")),
            "salarySnapshots": clean_amount(sum(scalar(row["total"]) for rows in all_payroll.values() for row in rows)),
        },
        "months": monthly_report,
        "warnings": [
            {
                "period": item["period"],
                "type": "payroll-reconciliation",
                "difference": item["payrollDelta"],
                "message": "薪資簿實領合計與營業額檔薪資支出欄不同；系統採薪資簿快照。",
            }
            for item in all_reconciliation if abs(scalar(item.get("payrollDelta"))) > 1
        ],
    }
    REPORT_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    REPORT_OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8", newline="\n")
    print(json.dumps(report["totals"], ensure_ascii=False, indent=2))
    print(f"months={len(monthly_report)} warnings={len(report['warnings'])}")


if __name__ == "__main__":
    main()
